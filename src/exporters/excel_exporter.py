"""
Excel Exporter for Stripe payout data
"""

import os
from datetime import datetime
from decimal import Decimal
from typing import List, Dict, Any, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from ..models import PayoutExportData, TransactionRecord, InvoiceRecord, FeeRecord, PayoutSummary
from ..utils import format_date_fr


class ExcelExporter:
    """Export payout data to Excel with French formatting and multiple sheets."""
    
    # Style definitions
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    TITLE_FONT = Font(bold=True, size=14)
    SUBTITLE_FONT = Font(bold=True, size=12)
    CURRENCY_FONT = Font(name="Calibri", size=11)
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self, output_dir: str):
        """
        Initialize Excel exporter.
        
        Args:
            output_dir: Directory to save Excel files
        """
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    def _format_decimal(self, value: Decimal) -> float:
        """Convert Decimal to float for Excel."""
        return float(value)
    
    def _apply_header_style(self, ws, row: int, cols: int):
        """Apply header styling to a row."""
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.THIN_BORDER
    
    def _apply_data_style(self, ws, start_row: int, end_row: int, cols: int):
        """Apply styling to data cells."""
        for row in range(start_row, end_row + 1):
            for col in range(1, cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = self.THIN_BORDER
                cell.alignment = Alignment(vertical='center')
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content."""
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    def _create_summary_sheet(self, wb: Workbook, summary: PayoutSummary):
        """Create the summary (Résumé) sheet."""
        ws = wb.active
        ws.title = "Résumé"
        
        # Title
        ws['A1'] = "RÉCAPITULATIF DU VIREMENT"
        ws['A1'].font = self.TITLE_FONT
        ws.merge_cells('A1:D1')
        
        # Payout info section
        ws['A3'] = "Informations du Virement"
        ws['A3'].font = self.SUBTITLE_FONT
        
        info_data = [
            ("ID Payout", summary.payout_id),
            ("Date", format_date_fr(summary.date)),
            ("Date Arrivée", format_date_fr(summary.date_arrivee)),
            ("Montant", self._format_decimal(summary.montant)),
            ("Devise", summary.devise),
            ("Statut", summary.statut),
            ("Méthode", summary.methode),
            ("Banque", summary.banque),
        ]
        
        row = 4
        for label, value in info_data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Financial breakdown section
        row += 1
        ws.cell(row=row, column=1, value="Détail Financier").font = self.SUBTITLE_FONT
        row += 1
        
        financial_data = [
            ("Total Paiements", self._format_decimal(summary.total_paiements), summary.devise),
            ("Total Remboursements", self._format_decimal(summary.total_remboursements), summary.devise),
            ("Total Frais Stripe", self._format_decimal(summary.total_frais), summary.devise),
            ("Total Litiges", self._format_decimal(summary.total_litiges), summary.devise),
            ("Total Autres", self._format_decimal(summary.total_autres), summary.devise),
        ]
        
        for label, amount, currency in financial_data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            cell = ws.cell(row=row, column=2, value=amount)
            cell.number_format = '#,##0.00'
            ws.cell(row=row, column=3, value=currency)
            row += 1
        
        # Statistics section
        row += 1
        ws.cell(row=row, column=1, value="Statistiques").font = self.SUBTITLE_FONT
        row += 1
        
        stats_data = [
            ("Nombre de Transactions", summary.nb_transactions),
            ("Nombre de Factures", summary.nb_factures),
            ("Nombre de Remboursements", summary.nb_remboursements),
            ("Nombre de Litiges", summary.nb_litiges),
        ]
        
        for label, value in stats_data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 10
    
    def _create_transactions_sheet(self, wb: Workbook, transactions: List[TransactionRecord]):
        """Create the transactions sheet."""
        ws = wb.create_sheet("Transactions")
        
        headers = [
            "Date",
            "Référence",
            "Type",
            "Description",
            "Montant Brut",
            "Frais",
            "Montant Net",
            "Devise",
            "Client",
            "N° Facture"
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._apply_header_style(ws, 1, len(headers))
        
        # Write data
        for row_num, tx in enumerate(transactions, 2):
            ws.cell(row=row_num, column=1, value=format_date_fr(tx.date))
            ws.cell(row=row_num, column=2, value=tx.reference)
            ws.cell(row=row_num, column=3, value=tx.type)
            ws.cell(row=row_num, column=4, value=tx.description)
            
            cell = ws.cell(row=row_num, column=5, value=self._format_decimal(tx.montant_brut))
            cell.number_format = '#,##0.00'
            
            cell = ws.cell(row=row_num, column=6, value=self._format_decimal(tx.frais))
            cell.number_format = '#,##0.00'
            
            cell = ws.cell(row=row_num, column=7, value=self._format_decimal(tx.montant_net))
            cell.number_format = '#,##0.00'
            
            ws.cell(row=row_num, column=8, value=tx.devise)
            ws.cell(row=row_num, column=9, value=tx.client or "")
            ws.cell(row=row_num, column=10, value=tx.numero_facture or "")
        
        self._apply_data_style(ws, 2, len(transactions) + 1, len(headers))
        self._auto_adjust_columns(ws)
    
    def _create_invoices_sheet(self, wb: Workbook, invoices: List[InvoiceRecord]):
        """Create the invoices (Factures) sheet."""
        if not invoices:
            return
        
        ws = wb.create_sheet("Factures")
        
        headers = [
            "N° Facture",
            "Date",
            "Date Échéance",
            "Client",
            "Email",
            "Montant HT",
            "TVA",
            "Montant TTC",
            "Devise",
            "Statut",
            "ID Stripe"
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._apply_header_style(ws, 1, len(headers))
        
        # Write data
        for row_num, inv in enumerate(invoices, 2):
            ws.cell(row=row_num, column=1, value=inv.numero)
            ws.cell(row=row_num, column=2, value=format_date_fr(inv.date))
            ws.cell(row=row_num, column=3, value=format_date_fr(inv.date_echeance) if inv.date_echeance else "")
            ws.cell(row=row_num, column=4, value=inv.client_nom)
            ws.cell(row=row_num, column=5, value=inv.client_email)
            
            cell = ws.cell(row=row_num, column=6, value=self._format_decimal(inv.montant_ht))
            cell.number_format = '#,##0.00'
            
            cell = ws.cell(row=row_num, column=7, value=self._format_decimal(inv.montant_tva))
            cell.number_format = '#,##0.00'
            
            cell = ws.cell(row=row_num, column=8, value=self._format_decimal(inv.montant_ttc))
            cell.number_format = '#,##0.00'
            
            ws.cell(row=row_num, column=9, value=inv.devise)
            ws.cell(row=row_num, column=10, value=inv.statut)
            ws.cell(row=row_num, column=11, value=inv.stripe_id)
        
        self._apply_data_style(ws, 2, len(invoices) + 1, len(headers))
        self._auto_adjust_columns(ws)
    
    def _create_fees_sheet(self, wb: Workbook, fees: List[FeeRecord]):
        """Create the fees (Frais) sheet."""
        if not fees:
            return
        
        ws = wb.create_sheet("Frais")
        
        headers = [
            "Transaction",
            "Type",
            "Description",
            "Montant",
            "Devise"
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._apply_header_style(ws, 1, len(headers))
        
        # Write data
        for row_num, fee in enumerate(fees, 2):
            ws.cell(row=row_num, column=1, value=fee.transaction_id)
            ws.cell(row=row_num, column=2, value=fee.type)
            ws.cell(row=row_num, column=3, value=fee.description)
            
            cell = ws.cell(row=row_num, column=4, value=self._format_decimal(fee.montant))
            cell.number_format = '#,##0.00'
            
            ws.cell(row=row_num, column=5, value=fee.devise)
        
        self._apply_data_style(ws, 2, len(fees) + 1, len(headers))
        self._auto_adjust_columns(ws)
        
        # Add total row
        total_row = len(fees) + 3
        ws.cell(row=total_row, column=3, value="TOTAL FRAIS").font = Font(bold=True)
        
        total_fees = sum(fee.montant for fee in fees)
        cell = ws.cell(row=total_row, column=4, value=self._format_decimal(total_fees))
        cell.number_format = '#,##0.00'
        cell.font = Font(bold=True)
        
        if fees:
            ws.cell(row=total_row, column=5, value=fees[0].devise)
    
    def export(self, data: PayoutExportData, filename: str = "recap_payout.xlsx") -> str:
        """
        Export all payout data to an Excel workbook.
        
        Args:
            data: PayoutExportData object
            filename: Output filename
            
        Returns:
            Path to the created file
        """
        filepath = os.path.join(self.output_dir, filename)
        
        wb = Workbook()
        
        # Create sheets
        self._create_summary_sheet(wb, data.summary)
        self._create_transactions_sheet(wb, data.transactions)
        self._create_invoices_sheet(wb, data.invoices)
        self._create_fees_sheet(wb, data.fees)
        
        wb.save(filepath)
        return filepath

